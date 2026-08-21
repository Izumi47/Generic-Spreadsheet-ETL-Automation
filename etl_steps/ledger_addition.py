"""Create a configurable ledger-entry workbook."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .common import (
    apply_template_styles,
    clear_data_rows,
    format_date,
    mapping_value,
    template_row_styles,
    text_value,
)


def _render(template, **values):
    try:
        return template.format(**values)
    except KeyError as error:
        raise ValueError(
            f"Ledger identifier template refers to unknown placeholder: "
            f"{error.args[0]}"
        ) from error


def _require_source_columns(source_df, source_columns):
    required = set(source_columns.values())
    missing = sorted(required.difference(source_df.columns))
    if missing:
        raise ValueError(f"Source workbook is missing columns: {missing}")


def run(source_df, mapping_df, template_path, output_path, config):
    """Write one ledger row for every eligible source record."""
    output_path = Path(output_path)
    if source_df.empty:
        return output_path, 0
    if mapping_df.empty:
        raise ValueError("The mapping workbook does not contain a mapping row")

    source_columns = config.get("source_columns", {})
    _require_source_columns(source_df, source_columns)
    mapping_row = mapping_df.iloc[0]

    workbook = load_workbook(template_path)
    sheet_name = config.get("template_sheet", "Ledger")
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Template does not contain a '{sheet_name}' worksheet")
    worksheet = workbook[sheet_name]

    expected_headers = config.get("expected_headers", [])
    header_count = len(expected_headers) or worksheet.max_column
    headers = [
        worksheet.cell(row=1, column=column).value
        for column in range(1, header_count + 1)
    ]
    if expected_headers and headers != expected_headers:
        workbook.close()
        raise ValueError(
            "Ledger template headers do not match the configured expected layout"
        )

    output_columns = config.get("output_columns", {})
    required_headers = set(output_columns.values())
    required_headers.update(config.get("mapping_assignments", {}).keys())
    required_headers.update(config.get("constants", {}).keys())
    missing_headers = sorted(required_headers.difference(headers))
    if missing_headers:
        workbook.close()
        raise ValueError(f"Ledger template is missing columns: {missing_headers}")

    styles = template_row_styles(worksheet)
    clear_data_rows(worksheet)
    generated_id_column = output_columns.get("generated_id", "SourceImportID")
    record_code_column = output_columns.get("record_code", "RecordCode")
    start_date_column = output_columns.get("start_date", "Period From")
    end_date_column = output_columns.get("end_date", "Period To")
    record_code_template = config.get(
        "record_code_template", "{entity_code}{counter}{record}"
    )
    generated_id_template = config.get(
        "generated_id_template", "{id_prefix}{record}"
    )
    entity_code = text_value(config.get("entity_code", ""))
    id_prefix = text_value(config.get("id_prefix", "Ledger_"))

    for counter, source_values in enumerate(
        source_df.itertuples(index=False, name=None), start=1
    ):
        source_record = dict(zip(source_df.columns, source_values))
        record = text_value(source_record.get(source_columns.get("record", ""), ""))
        values = {header: "" for header in headers}
        values[generated_id_column] = _render(
            generated_id_template,
            id_prefix=id_prefix,
            entity_code=entity_code,
            counter=counter,
            record=record,
        )
        values[record_code_column] = _render(
            record_code_template,
            entity_code=entity_code,
            counter=counter,
            record=record,
        )
        values[start_date_column] = format_date(
            source_record.get(source_columns.get("start_date", ""), "")
        )
        values[end_date_column] = format_date(
            source_record.get(source_columns.get("end_date", ""), "")
        )

        for output_column, mapping_column in config.get(
            "mapping_assignments", {}
        ).items():
            values[output_column] = mapping_value(mapping_row, mapping_column)
        values.update(config.get("constants", {}))

        row_number = counter + 1
        apply_template_styles(worksheet, row_number, styles)
        for header, value in values.items():
            column_number = headers.index(header) + 1
            worksheet.cell(row=row_number, column=column_number, value=value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return output_path, len(source_df)
