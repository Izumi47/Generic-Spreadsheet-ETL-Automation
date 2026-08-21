"""Create a configurable allocation workbook from lookup values."""

from pathlib import Path
import re
import shutil

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from .common import (
    apply_template_styles,
    clear_data_rows,
    is_missing,
    mapping_value,
    template_row_styles,
    text_value,
)


class LookupValidationError(ValueError):
    """Raised when a configured lookup rule cannot resolve a source value."""

    def __init__(self, failures):
        self.failures = failures
        summary = "\n".join(
            f"  Source row {failure['source_row']}: "
            f"{failure['output_column']} not found for "
            f"'{failure['search_text']}'"
            for failure in failures
        )
        super().__init__(f"{len(failures)} lookup failure(s):\n{summary}")


def _code_value(value) -> str:
    if is_missing(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return text_value(value)


def _render(template, **values):
    try:
        return template.format(**values)
    except KeyError as error:
        raise ValueError(
            f"Allocation identifier template refers to unknown placeholder: "
            f"{error.args[0]}"
        ) from error


def lookup_option(
    lookup_df,
    column,
    search_text,
    *,
    contains=False,
    right_side=False,
    values_cache=None,
    result_cache=None,
):
    """Return the first lookup value matching the configured search rule."""
    if column not in lookup_df or is_missing(search_text):
        return ""

    search_text = text_value(search_text)
    if not search_text:
        return ""
    cache_key = (column, search_text, contains, right_side)
    if result_cache is not None and cache_key in result_cache:
        return result_cache[cache_key]

    if values_cache is None:
        values = lookup_df[column].dropna().astype(str).str.strip()
    else:
        values = values_cache.setdefault(
            column,
            lookup_df[column].dropna().astype(str).str.strip(),
        )

    needle = f"({search_text})" if right_side else search_text
    if contains:
        matches = values[values.str.contains(re.escape(needle), na=False)]
    else:
        matches = values[values == needle]
    result = matches.iloc[0] if not matches.empty else ""
    if result_cache is not None:
        result_cache[cache_key] = result
    return result


def _validate_configuration(source_df, dest_df, lookup_df, config):
    source_columns = config.get("source_columns", {})
    missing_source = sorted(
        set(source_columns.values()).difference(source_df.columns)
    )
    if missing_source:
        raise ValueError(f"Source workbook is missing columns: {missing_source}")

    required_output = set(config.get("output_columns", {}).values())
    required_output.update(
        rule.get("output_column") for rule in config.get("lookup_rules", [])
    )
    required_output.update(config.get("default_values", {}).keys())
    missing_output = sorted(required_output.difference(dest_df.columns))
    if missing_output:
        raise ValueError(f"Allocation template is missing columns: {missing_output}")

    for rule in config.get("lookup_rules", []):
        lookup_column = rule.get("lookup_column")
        if lookup_df is None or lookup_column not in lookup_df.columns:
            raise ValueError(
                f"Lookup workbook is missing column '{lookup_column}'"
            )


def allocation_data(source_df, dest_df, mapping_df, lookup_df, config):
    """Build the allocation DataFrame and collect all lookup failures."""
    if source_df.empty:
        return pd.DataFrame(columns=dest_df.columns), []
    if mapping_df.empty:
        raise ValueError("The mapping workbook does not contain a mapping row")
    if lookup_df is None:
        raise ValueError("Allocation lookup rules require a lookup worksheet")

    _validate_configuration(source_df, dest_df, lookup_df, config)
    source_columns = config.get("source_columns", {})
    output_columns = config.get("output_columns", {})
    id_column = output_columns.get("id", "SourceImportID")
    id_template = config.get("id_template", "{entity_code}{counter}{record}")
    entity_code = _code_value(config.get("entity_code", ""))
    lookup_values_cache = {}
    lookup_result_cache = {}
    output = []
    failures = []

    for counter, source_values in enumerate(
        source_df.itertuples(index=False, name=None), start=1
    ):
        source_record = dict(zip(source_df.columns, source_values))
        output_row = {column: None for column in dest_df.columns}
        record = text_value(source_record.get(source_columns.get("record", ""), ""))
        output_row[id_column] = _render(
            id_template,
            entity_code=entity_code,
            counter=counter,
            record=record,
        )

        for rule in config.get("lookup_rules", []):
            output_column = rule["output_column"]
            search_values = [
                text_value(source_record.get(column, ""))
                for column in rule.get("source_columns", ())
            ]
            search_text = rule.get("joiner", "").join(search_values)
            result = lookup_option(
                lookup_df,
                rule["lookup_column"],
                search_text,
                contains=bool(rule.get("contains", False)),
                right_side=bool(rule.get("right_side", False)),
                values_cache=lookup_values_cache,
                result_cache=lookup_result_cache,
            )
            output_row[output_column] = result
            if not result:
                source_row = (
                    source_df.index[counter - 1] + 2
                    if isinstance(source_df.index[counter - 1], int)
                    else counter + 1
                )
                failures.append(
                    {
                        "source_row": source_row,
                        "output_column": output_column,
                        "search_text": search_text,
                        "record": record,
                        "entity_code": entity_code,
                    }
                )

        for column, value in config.get("default_values", {}).items():
            if column in output_row:
                output_row[column] = value
        output.append(output_row)

    return pd.DataFrame(output, columns=dest_df.columns), failures


def add_lookup_dropdowns(
    workbook,
    worksheet,
    lookup_df,
    lookup_sheet_name,
    dropdown_columns,
    first_data_row=2,
):
    """Copy lookup values into the output and attach Excel list validations."""
    if not dropdown_columns:
        return
    if lookup_sheet_name in workbook.sheetnames:
        del workbook[lookup_sheet_name]
    lookup_sheet = workbook.create_sheet(lookup_sheet_name)
    lookup_sheet.sheet_state = "visible"

    lookup_columns = list(lookup_df.columns)
    for column_number, column_name in enumerate(lookup_columns, start=1):
        lookup_sheet.cell(row=1, column=column_number, value=column_name)
        for row_number, value in enumerate(lookup_df[column_name], start=2):
            lookup_sheet.cell(
                row=row_number,
                column=column_number,
                value=None if is_missing(value) else value,
            )

    last_lookup_row = max(2, len(lookup_df) + 1)
    last_output_row = max(first_data_row, worksheet.max_row)
    escaped_sheet_name = lookup_sheet_name.replace("'", "''")

    for column_name in dropdown_columns:
        if column_name not in lookup_columns:
            raise ValueError(
                f"Lookup workbook is missing dropdown column '{column_name}'"
            )
        target_cell = next(
            (cell for cell in worksheet[1] if cell.value == column_name),
            None,
        )
        if target_cell is None:
            raise ValueError(
                f"Output worksheet is missing dropdown column '{column_name}'"
            )

        lookup_column = get_column_letter(lookup_columns.index(column_name) + 1)
        validation = DataValidation(
            type="list",
            formula1=(
                f'=INDIRECT("\'{escaped_sheet_name}\'!${lookup_column}$'
                f'2:${lookup_column}${last_lookup_row}")'
            ),
            allow_blank=True,
        )
        validation.error = f"Select a value from the {column_name} list."
        validation.errorTitle = "Invalid value"
        validation.prompt = f"Select a {column_name} value."
        validation.promptTitle = "List value"
        worksheet.add_data_validation(validation)
        target_column = get_column_letter(target_cell.column)
        validation.add(
            f"{target_column}{first_data_row}:{target_column}{last_output_row}"
        )


def save_allocation_workbook(
    result_df,
    template_path,
    output_path,
    config,
    lookup_df=None,
):
    """Save allocation rows while retaining the configured template workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    workbook = load_workbook(output_path)
    sheet_name = config.get("template_sheet", "Allocations")
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Template does not contain a '{sheet_name}' worksheet")
    worksheet = workbook[sheet_name]
    styles = template_row_styles(worksheet)
    clear_data_rows(worksheet)

    for column_number, column_name in enumerate(result_df.columns, start=1):
        worksheet.cell(row=1, column=column_number, value=column_name)
    for row_number, row in enumerate(
        result_df.itertuples(index=False, name=None), start=2
    ):
        apply_template_styles(worksheet, row_number, styles)
        for column_number, value in enumerate(row, start=1):
            worksheet.cell(
                row=row_number,
                column=column_number,
                value=None if is_missing(value) else value,
            )

    if lookup_df is not None:
        add_lookup_dropdowns(
            workbook,
            worksheet,
            lookup_df,
            config.get("lookup_output_sheet", "Lookup Values"),
            config.get("dropdown_columns", ()),
        )

    workbook.save(output_path)
    workbook.close()
    return output_path


def run(source_df, mapping_df, dest_df, lookup_df, template_path, output_path, config):
    """Build, validate, and save the allocation output."""
    output_path = Path(output_path)
    result_df, failures = allocation_data(
        source_df, dest_df, mapping_df, lookup_df, config
    )
    if failures:
        raise LookupValidationError(failures)
    if result_df.empty:
        return output_path, 0
    save_allocation_workbook(
        result_df,
        template_path,
        output_path,
        config,
        lookup_df=lookup_df,
    )
    return output_path, len(result_df)
