"""Create a configurable charge or transaction addition workbook."""

from copy import copy
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.datetime import to_excel

from .common import (
    apply_template_styles,
    clear_data_rows,
    format_date,
    is_missing,
    mapping_value,
    template_row_styles,
    text_value,
)


def _format_code(value) -> str:
    if is_missing(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return text_value(value)


def _render(template, **values):
    try:
        return template.format(**values)
    except KeyError as error:
        raise ValueError(
            f"Charge identifier template refers to unknown placeholder: "
            f"{error.args[0]}"
        ) from error


def _date_serial(value) -> str:
    if is_missing(value) or not isinstance(value, (date, datetime, pd.Timestamp)):
        return ""
    formatted = format_date(value)
    return str(int(to_excel(formatted)))


def _require_source_columns(source_df, source_columns):
    required = set(source_columns.values())
    missing = sorted(required.difference(source_df.columns))
    if missing:
        raise ValueError(f"Source workbook is missing columns: {missing}")


def run(source_df, mapping_df, template_path, output_path, config):
    """Write one configurable charge row for every eligible source record."""
    output_path = Path(output_path)
    if source_df.empty:
        return output_path, 0
    if mapping_df.empty:
        raise ValueError("The mapping workbook does not contain a mapping row")

    source_columns = config.get("source_columns", {})
    _require_source_columns(source_df, source_columns)
    mapping_row = mapping_df.iloc[0]

    workbook = load_workbook(template_path)
    sheet_name = config.get("template_sheet", "Charges")
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Template does not contain a '{sheet_name}' worksheet")
    worksheet = workbook[sheet_name]
    styles = template_row_styles(worksheet)
    clear_data_rows(worksheet)

    output_columns = config.get("output_columns", {})
    generated_id_column = output_columns.get("generated_id", "A")
    source_id_column = output_columns.get("source_id", "B")
    source_assignments = config.get("source_assignments", {})
    mapping_assignments = config.get("mapping_assignments", {})
    prefix_column = config.get("mapping_prefix_column", "")
    source_id_template = config.get(
        "source_id_template", "{prefix}{counter}{reference}"
    )
    generated_id_template = config.get(
        "generated_id_template", "{source_id}{start_serial}{amount}"
    )
    entity_code = text_value(config.get("entity_code", ""))
    identifier_fill = PatternFill(
        fill_type="solid",
        fgColor=config.get("identifier_fill", "4F81BD"),
    )

    for counter, source_values in enumerate(
        source_df.itertuples(index=False, name=None), start=1
    ):
        source_record = dict(zip(source_df.columns, source_values))
        reference = text_value(
            source_record.get(source_columns.get("reference", ""), "")
        )
        start_date = format_date(
            source_record.get(source_columns.get("start_date", ""), "")
        )
        amount_value = source_record.get(source_columns.get("amount", ""), "")
        amount = _format_code(amount_value)
        prefix = text_value(mapping_value(mapping_row, prefix_column))
        source_id = _render(
            source_id_template,
            prefix=prefix,
            counter=counter,
            reference=reference,
            entity_code=entity_code,
        )
        generated_id = _render(
            generated_id_template,
            source_id=source_id,
            start_serial=_date_serial(start_date),
            amount=amount,
            reference=reference,
            counter=counter,
            entity_code=entity_code,
        )

        values = {
            generated_id_column: generated_id,
            source_id_column: source_id,
        }
        for column, source_key in source_assignments.items():
            source_column = source_columns.get(source_key)
            if source_column is None:
                raise ValueError(
                    f"Charge source assignment '{source_key}' is not configured"
                )
            value = source_record.get(source_column, "")
            values[column] = format_date(value)
        for column, mapping_column in mapping_assignments.items():
            values[column] = mapping_value(mapping_row, mapping_column)

        row_number = counter + 1
        apply_template_styles(worksheet, row_number, styles)
        for column, value in values.items():
            worksheet[f"{column}{row_number}"] = value

        identifier_cell = worksheet[f"{generated_id_column}{row_number}"]
        identifier_cell.fill = identifier_fill
        font = copy(identifier_cell.font)
        font.color = "FFFFFF"
        identifier_cell.font = font

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return output_path, len(source_df)
