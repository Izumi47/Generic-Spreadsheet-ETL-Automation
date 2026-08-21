"""Create a configurable record-addition workbook."""

from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .common import (
    apply_template_styles,
    clear_data_rows,
    format_date,
    mapping_value,
    normalise_boolean,
    template_row_styles,
    text_value,
)


def _require_source_columns(source_df, source_columns):
    required = set(source_columns.values())
    missing = sorted(required.difference(source_df.columns))
    if missing:
        raise ValueError(f"Source workbook is missing columns: {missing}")


def _render(template, **values):
    try:
        return template.format(**values)
    except KeyError as error:
        raise ValueError(
            f"Identifier template refers to unknown placeholder: {error.args[0]}"
        ) from error


def run(source_df, mapping_df, template_path, output_path, config):
    """Write one output row for every eligible source record."""
    output_path = Path(output_path)
    if source_df.empty:
        return output_path, 0
    if mapping_df.empty:
        raise ValueError("The mapping workbook does not contain a mapping row")

    source_columns = config.get("source_columns", {})
    _require_source_columns(source_df, source_columns)
    mapping_row = mapping_df.iloc[0]

    workbook = load_workbook(template_path)
    sheet_name = config.get("template_sheet", "Records")
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Template does not contain a '{sheet_name}' worksheet")
    worksheet = workbook[sheet_name]
    styles = template_row_styles(worksheet)
    clear_data_rows(worksheet)

    output_columns = config.get("output_columns", {})
    mapping_assignments = config.get("mapping_assignments", {})
    source_assignments = config.get("source_assignments", {})
    prefix_column = config.get("mapping_prefix_column", "")
    id_template = config.get("id_template", "{entity_code}{counter}{reference}")
    record_code_template = config.get(
        "record_code_template", "{entity_code}_{identifier_prefix}"
    )
    entity_code = text_value(config.get("entity_code", ""))
    identifier_prefix = text_value(mapping_value(mapping_row, prefix_column))
    boolean_columns = set(config.get("boolean_output_columns", ()))
    identifier_column = output_columns.get("id", "A")
    record_code_column = output_columns.get("record_code", "B")
    reference_key = source_columns.get("reference")

    identifier_fill = PatternFill(
        fill_type="solid",
        fgColor=config.get("identifier_fill", "4F81BD"),
    )

    for counter, source_values in enumerate(
        source_df.itertuples(index=False, name=None), start=1
    ):
        source_record = dict(zip(source_df.columns, source_values))
        reference = text_value(source_record.get(reference_key, ""))
        values = {
            column: mapping_value(mapping_row, mapping_column)
            for column, mapping_column in mapping_assignments.items()
        }

        for column, source_key in source_assignments.items():
            source_column = source_columns.get(source_key)
            if source_column is None:
                raise ValueError(
                    f"Record source assignment '{source_key}' is not configured"
                )
            value = source_record.get(source_column, "")
            values[column] = format_date(value)

        values[identifier_column] = _render(
            id_template,
            entity_code=entity_code,
            counter=counter,
            reference=reference,
            identifier_prefix=identifier_prefix,
        )
        values[record_code_column] = _render(
            record_code_template,
            entity_code=entity_code,
            counter=counter,
            reference=reference,
            identifier_prefix=identifier_prefix,
        )

        for column in boolean_columns:
            if column in values:
                values[column] = normalise_boolean(values[column])

        row_number = counter + 1
        apply_template_styles(worksheet, row_number, styles)
        for column, value in values.items():
            worksheet[f"{column}{row_number}"] = value

        identifier_cell = worksheet[f"{identifier_column}{row_number}"]
        identifier_cell.fill = identifier_fill
        font = copy(identifier_cell.font)
        font.color = "FFFFFF"
        identifier_cell.font = font

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return output_path, len(source_df)
