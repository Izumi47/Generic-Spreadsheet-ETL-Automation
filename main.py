"""Run a configurable, non-domain-specific spreadsheet ETL workflow.

Copy config.example.json to config.json, replace its placeholders with local
workbook names and column mappings, then run:

    python main.py

The controller deliberately contains no organization names, URLs, entity
identifiers, or business-system labels. All workbook-specific values belong in
configuration or in the supplied workbooks.
"""

from __future__ import annotations

import argparse
from copy import copy
from datetime import date, datetime
import json
import logging
from numbers import Integral
from pathlib import Path
import re
import sys
from time import perf_counter

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from etl_steps import allocation_addition as allocation
from etl_steps import charge_addition as charge
from etl_steps import ledger_addition as ledger
from etl_steps import record_addition as record


BASE_DIR = (
    Path(sys.executable).resolve().parent
    if getattr(sys, "frozen", False)
    else Path(__file__).resolve().parent
)
DEFAULT_CONFIG_PATH = BASE_DIR / "config.json"
EXAMPLE_CONFIG_PATH = BASE_DIR / "config.example.json"
LOGGER = logging.getLogger("generic_spreadsheet_etl")

STEP_RUNNERS = {
    "allocation": allocation,
    "record": record,
    "charge": charge,
    "ledger": ledger,
}


def _resolve_path(value, base_dir=BASE_DIR):
    path = Path(str(value))
    return path if path.is_absolute() else base_dir / path


def _text(value) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalised_series(series):
    return series.map(_text)


def _safe_path_part(value) -> str:
    """Keep user-provided labels from becoming unsafe path components."""
    cleaned = re.sub(r"[^A-Za-z0-9._ -]+", "_", _text(value)).strip(" .")
    return cleaned or "unknown"


def load_config(config_path=None):
    """Load config.json, falling back to the non-confidential example config."""
    requested_path = _resolve_path(config_path or DEFAULT_CONFIG_PATH)
    if not requested_path.is_file():
        if requested_path == DEFAULT_CONFIG_PATH and EXAMPLE_CONFIG_PATH.is_file():
            requested_path = EXAMPLE_CONFIG_PATH
        else:
            raise FileNotFoundError(
                f"Configuration file not found: {requested_path}. "
                "Copy config.example.json to config.json and edit the placeholders."
            )

    with requested_path.open("r", encoding="utf-8") as stream:
        config = json.load(stream)
    if not isinstance(config, dict):
        raise ValueError("The configuration file must contain a JSON object")
    LOGGER.info("Configuration loaded from %s", requested_path.name)
    return config


def configure_logging(output_root):
    """Create one plain-text audit log for the current run."""
    output_root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = output_root / f"ETL_Audit_{timestamp}.log"

    for handler in LOGGER.handlers[:]:
        handler.close()
        LOGGER.removeHandler(handler)
    LOGGER.setLevel(logging.INFO)
    LOGGER.propagate = False

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(
        logging.Formatter(
            "%(asctime)s | %(levelname)s | %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(logging.Formatter("%(message)s"))
    LOGGER.addHandler(file_handler)
    LOGGER.addHandler(console_handler)
    LOGGER.info("Audit log started: %s", log_path.name)
    return log_path


def mark_failed_log(log_path):
    """Rename a completed audit log so failed runs are easy to identify."""
    if log_path is None or not log_path.is_file():
        return
    for handler in LOGGER.handlers[:]:
        handler.flush()
        handler.close()
        LOGGER.removeHandler(handler)
    error_path = log_path.with_name(f"ERROR_{log_path.name}")
    log_path.replace(error_path)


def validate_inputs(config):
    """Resolve and validate the configured workbooks and worksheets."""
    application = config.get("application", {})
    source_path = _resolve_path(application.get("source_file", ""))
    mapping_path = _resolve_path(application.get("mapping_file", ""))
    template_directory = _resolve_path(
        application.get("template_directory", "Required Data/Empty Template File")
    )

    required_files = [source_path, mapping_path]
    missing = [path for path in required_files if not path.is_file()]
    for step_id in config.get("step_order", ()):
        step = config.get("steps", {}).get(step_id, {})
        template_path = template_directory / str(step.get("template_file", ""))
        required_files.append(template_path)
        if not template_path.is_file():
            missing.append(template_path)

    if missing:
        missing_list = "\n".join(f"  - {path}" for path in dict.fromkeys(missing))
        raise FileNotFoundError(
            "Required workbook(s) are missing. Replace the placeholders in "
            f"config.json and add the workbooks:\n{missing_list}"
        )

    return {
        "source_path": source_path,
        "mapping_path": mapping_path,
        "template_directory": template_directory,
    }


def load_mapping(mapping_path, mapping_sheet):
    mapping_df = pd.read_excel(mapping_path, sheet_name=mapping_sheet).dropna(
        how="all"
    )
    if mapping_df.empty:
        raise ValueError("The mapping workbook does not contain any data rows")
    return mapping_df


def choose_entity_code(mapping_df, application, requested_entity=None):
    """Use an explicit entity code or select one from the mapping workbook."""
    entity_column = application.get("entity_column", "Entity Code")
    if entity_column not in mapping_df.columns:
        raise ValueError(
            f"Mapping worksheet is missing the entity column '{entity_column}'"
        )

    codes = sorted(
        {
            _text(value)
            for value in mapping_df[entity_column]
            if _text(value)
        }
    )
    configured_code = _text(
        requested_entity if requested_entity is not None else application.get("entity_code", "")
    )
    if configured_code:
        if configured_code not in codes:
            raise ValueError(
                f"Entity code '{configured_code}' was not found in the mapping workbook"
            )
        return configured_code
    if len(codes) == 1:
        return codes[0]
    if not codes:
        raise ValueError("The mapping worksheet does not contain entity codes")

    print("Available entity codes:")
    for number, code in enumerate(codes, start=1):
        print(f"  {number}. {code}")
    try:
        choice = input("Select an entity code by number: ").strip()
        selected_number = int(choice)
        return codes[selected_number - 1]
    except (EOFError, ValueError, IndexError) as error:
        raise RuntimeError(
            "An entity code is required. Set application.entity_code in "
            "config.json or pass --entity."
        ) from error


def filter_source_rows(source_df, application, processing, entity_code):
    """Apply shared source filtering before any output workbook is created."""
    if source_df.empty:
        raise ValueError("The source worksheet contains no data rows")

    required_columns = set(processing.get("required_source_columns", ()))
    missing = sorted(required_columns.difference(source_df.columns))
    if missing:
        raise ValueError(f"Source workbook is missing columns: {missing}")

    trim_count = int(processing.get("trim_trailing_rows", 0) or 0)
    if trim_count:
        source_df = (
            source_df.iloc[:-trim_count].copy()
            if len(source_df) > trim_count
            else source_df.iloc[0:0].copy()
        )

    filter_column = processing.get("filter_column")
    if filter_column:
        if filter_column not in source_df.columns:
            raise ValueError(
                f"Source workbook is missing the filter column '{filter_column}'"
            )
        filter_values = _normalised_series(source_df[filter_column])
        if processing.get("filter_blank_only", True):
            filter_mask = filter_values.eq("")
        else:
            allowed_values = {
                _text(value) for value in processing.get("eligible_filter_values", ())
            }
            filter_mask = filter_values.isin(allowed_values)
        filtered_rows = source_df.loc[filter_mask].copy()
    else:
        filtered_rows = source_df.copy()
    filter_excluded_rows = len(source_df) - len(filtered_rows)

    entity_excluded_rows = 0
    entity_column = application.get("entity_column", "Entity Code")
    if application.get("filter_source_to_entity", True):
        if entity_column not in filtered_rows.columns:
            raise ValueError(
                f"Source workbook is missing the entity column '{entity_column}'"
            )
        entity_mask = _normalised_series(filtered_rows[entity_column]).eq(
            _text(entity_code)
        )
        entity_excluded_rows = len(filtered_rows) - int(entity_mask.sum())
        filtered_rows = filtered_rows.loc[entity_mask].copy()

    first_column = filtered_rows.columns[0]
    excluded_values = {
        _text(value)
        for value in processing.get("excluded_first_column_values", ())
    }
    first_column_values = _normalised_series(filtered_rows[first_column])
    first_column_mask = first_column_values.ne("") & ~first_column_values.isin(
        excluded_values
    )
    first_column_excluded_rows = len(filtered_rows) - int(first_column_mask.sum())
    filtered_rows = filtered_rows.loc[first_column_mask].copy()

    date_excluded_rows = 0
    date_window = processing.get("date_window", {})
    if date_window.get("enabled", False):
        start_column = date_window.get("start_column")
        end_column = date_window.get("end_column")
        missing_dates = {
            column
            for column in (start_column, end_column)
            if column not in filtered_rows.columns
        }
        if missing_dates:
            raise ValueError(
                f"Source workbook is missing date-window columns: "
                f"{sorted(missing_dates)}"
            )
        if date_window.get("mode", "active_month") != "active_month":
            raise ValueError(
                "Only the generic 'active_month' date-window mode is supported"
            )
        current_month = pd.Timestamp(datetime.now().year, datetime.now().month, 1)
        next_month = current_month + pd.offsets.MonthBegin(1)
        start_dates = pd.to_datetime(
            filtered_rows[start_column], errors="coerce"
        )
        end_dates = pd.to_datetime(filtered_rows[end_column], errors="coerce")
        date_mask = (start_dates < next_month) & (end_dates >= current_month)
        date_excluded_rows = len(filtered_rows) - int(date_mask.sum())
        filtered_rows = filtered_rows.loc[date_mask].copy()

    summary = {
        "source_rows": len(source_df),
        "eligible_rows": len(filtered_rows),
        "filter_excluded_rows": filter_excluded_rows,
        "entity_excluded_rows": entity_excluded_rows,
        "first_column_excluded_rows": first_column_excluded_rows,
        "date_excluded_rows": date_excluded_rows,
        "filter_column": filter_column or "disabled",
        "entity_code": entity_code,
    }
    return filtered_rows, summary


def find_incomplete_rows(source_df, eligible_rows, processing):
    """Return a report DataFrame for rows with blank required fields."""
    if not processing.get("validate_complete_rows", True):
        return pd.DataFrame()

    configured_columns = processing.get("completeness_columns")
    columns_to_check = list(
        configured_columns
        or processing.get("required_source_columns")
        or source_df.columns
    )
    excluded_columns = set(processing.get("completeness_exclude_columns", ()))
    filter_column = processing.get("filter_column")
    if filter_column:
        excluded_columns.add(filter_column)
    columns_to_check = [
        column for column in columns_to_check if column not in excluded_columns
    ]
    missing_columns = sorted(set(columns_to_check).difference(source_df.columns))
    if missing_columns:
        raise ValueError(
            f"Completeness configuration refers to missing columns: {missing_columns}"
        )

    report_rows = []
    for position, (index, source_row) in enumerate(eligible_rows.iterrows(), start=2):
        missing = [
            column
            for column in columns_to_check
            if pd.isna(source_row[column]) or _text(source_row[column]) == ""
        ]
        if missing:
            report_row = source_row.to_dict()
            report_row["Source Row"] = int(index) + 2 if isinstance(index, Integral) else position
            report_row["Missing Columns"] = ", ".join(missing)
            report_rows.append(report_row)

    if not report_rows:
        return pd.DataFrame()
    report = pd.DataFrame(report_rows)
    return report.reindex(
        columns=["Source Row", "Missing Columns"] + list(source_df.columns)
    )


def output_folder(output_root, output_config, entity_code, now):
    """Build a safe dated output folder from a configuration template."""
    values = {
        "year": now.strftime("%Y"),
        "month_name": now.strftime("%m. %b"),
        "month_tag": now.strftime(output_config.get("month_tag_format", "%m%y")),
        "entity_code": _safe_path_part(entity_code),
    }
    template = output_config.get(
        "folder_template", "{year}/{month_name}/{entity_code}"
    )
    try:
        relative_folder = Path(template.format(**values))
    except KeyError as error:
        raise ValueError(
            f"Output folder template refers to unknown placeholder: {error.args[0]}"
        ) from error
    if relative_folder.is_absolute() or ".." in relative_folder.parts:
        raise ValueError("Output folder template must stay below the output directory")
    return output_root / relative_folder


def output_path(output_folder_path, step_config, output_config, now):
    values = {
        "number": int(step_config.get("number", 0)),
        "step_name": _safe_path_part(step_config.get("name", "step")),
        "month_tag": now.strftime(output_config.get("month_tag_format", "%m%y")),
    }
    template = output_config.get(
        "filename_template", "{number:02d}_{step_name}_{month_tag}.xlsx"
    )
    try:
        filename = template.format(**values)
    except KeyError as error:
        raise ValueError(
            f"Output filename template refers to unknown placeholder: {error.args[0]}"
        ) from error
    return output_folder_path / _safe_path_part(filename)


def normalise_output_fonts(workbook_path, output_config, last_data_row):
    """Apply consistent data-row formatting without changing template headers."""
    workbook = load_workbook(workbook_path)
    configured_date_columns = {
        _text(value).casefold() for value in output_config.get("date_columns", ())
    }
    for worksheet in workbook.worksheets:
        date_columns = set()
        for cell in worksheet[1]:
            header = _text(cell.value).casefold()
            if (
                header in configured_date_columns
                or "date" in header
                or header in {"period from", "period to"}
            ):
                date_columns.add(cell.column)

        data_end_row = min(last_data_row, worksheet.max_row)
        for row in worksheet.iter_rows(min_row=2, max_row=data_end_row):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if isinstance(cell.value, (date, datetime)) or cell.is_date:
                    date_columns.add(cell.column)

        for row in worksheet.iter_rows(min_row=2, max_row=data_end_row):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                font = copy(cell.font)
                font.name = "Calibri"
                font.sz = 11
                font.bold = False
                cell.font = font
                if cell.column in date_columns:
                    cell.number_format = "d/m/yyyy"
    workbook.save(workbook_path)
    workbook.close()


def write_incomplete_report(report, output_folder_path, month_tag):
    report_path = output_folder_path / f"Incomplete_Source_Rows_{month_tag}.xlsx"
    report.to_excel(report_path, index=False)
    LOGGER.warning(
        "Incomplete source data found: %d row(s); report saved to %s",
        len(report),
        report_path,
    )
    return report_path


def write_lookup_report(failures, output_folder_path, month_tag):
    report_path = output_folder_path / f"Lookup_Failures_{month_tag}.xlsx"
    pd.DataFrame(failures).to_excel(report_path, index=False)
    LOGGER.warning(
        "Lookup failures found: %d; report saved to %s",
        len(failures),
        report_path,
    )
    return report_path


def write_run_summary(summary, output_folder_path):
    summary_path = output_folder_path / "ETL_Run_Summary.json"
    with summary_path.open("w", encoding="utf-8") as stream:
        json.dump(summary, stream, indent=2, default=str)
    return summary_path


def run_step(
    step_id,
    step_config,
    source_df,
    mapping_df,
    template_path,
    output_path_value,
):
    """Dispatch one configured step while keeping shared data in memory."""
    module = STEP_RUNNERS.get(step_id)
    if module is None:
        raise ValueError(f"Unknown ETL step '{step_id}'")

    if step_id == "allocation":
        template_sheet = step_config.get("template_sheet", "Allocations")
        dest_df = pd.read_excel(template_path, sheet_name=template_sheet)
        lookup_df = pd.read_excel(
            template_path,
            sheet_name=step_config.get("lookup_sheet", "Lookup Values"),
            header=int(step_config.get("lookup_header_row", 0)),
        )
        return module.run(
            source_df,
            mapping_df,
            dest_df,
            lookup_df,
            template_path,
            output_path_value,
            step_config,
        )

    return module.run(
        source_df,
        mapping_df,
        template_path,
        output_path_value,
        step_config,
    )


def main(config_path=None, requested_entity=None):
    """Run all configured steps and return a JSON-serialisable summary."""
    config = load_config(config_path)
    application = config.get("application", {})
    processing = config.get("processing", {})
    output_config = config.get("output", {})
    output_root = _resolve_path(application.get("output_directory", "Output"))
    log_path = configure_logging(output_root)
    run_started_at = perf_counter()
    LOGGER.info("%s started", application.get("name", "Spreadsheet ETL"))

    try:
        paths = validate_inputs(config)
        mapping_sheet = application.get("mapping_sheet", "Mapping")
        source_sheet = application.get("source_sheet", "SourceData")
        mapping_df = load_mapping(paths["mapping_path"], mapping_sheet)
        entity_code = choose_entity_code(
            mapping_df,
            application,
            requested_entity=requested_entity,
        )
        selected_mapping = mapping_df.loc[
            _normalised_series(mapping_df[application.get("entity_column", "Entity Code")]).eq(
                entity_code
            )
        ].copy()
        if selected_mapping.empty:
            raise ValueError(f"No mapping row exists for entity code '{entity_code}'")

        source_df = pd.read_excel(paths["source_path"], sheet_name=source_sheet).copy()
        eligible_rows, run_summary = filter_source_rows(
            source_df,
            application,
            processing,
            entity_code,
        )
        if eligible_rows.empty:
            raise ValueError("No source rows remain after the configured filters")

        incomplete_report = find_incomplete_rows(
            source_df,
            eligible_rows,
            processing,
        )
        now = datetime.now()
        run_output_folder = output_folder(
            output_root,
            output_config,
            entity_code,
            now,
        )
        run_output_folder.mkdir(parents=True, exist_ok=True)
        month_tag = now.strftime(output_config.get("month_tag_format", "%m%y"))
        if not incomplete_report.empty:
            write_incomplete_report(
                incomplete_report,
                run_output_folder,
                month_tag,
            )
            if processing.get("stop_on_incomplete", True):
                raise ValueError(
                    f"{len(incomplete_report)} eligible source row(s) contain "
                    "missing data"
                )

        LOGGER.info("Selected entity: %s", entity_code)
        LOGGER.info("Source: %s", paths["source_path"].name)
        LOGGER.info("Mapping: %s", paths["mapping_path"].name)
        LOGGER.info("Output folder: %s", run_output_folder)
        LOGGER.info("Eligible source rows: %d", len(eligible_rows))

        step_results = []
        for step_id in config.get("step_order", ()):
            step_config = dict(config.get("steps", {}).get(step_id, {}))
            if not step_config:
                raise ValueError(f"No configuration exists for ETL step '{step_id}'")
            step_config["entity_code"] = entity_code
            template_path = (
                paths["template_directory"] / step_config["template_file"]
            )
            step_output_path = output_path(
                run_output_folder,
                step_config,
                output_config,
                now,
            )
            LOGGER.info(
                "[%d/%d] Starting %s",
                int(step_config.get("number", len(step_results) + 1)),
                len(config.get("step_order", ())),
                step_config.get("name", step_id),
            )
            step_started_at = perf_counter()
            try:
                output_file, row_count = run_step(
                    step_id,
                    step_config,
                    eligible_rows,
                    selected_mapping,
                    template_path,
                    step_output_path,
                )
            except allocation.LookupValidationError as error:
                write_lookup_report(error.failures, run_output_folder, month_tag)
                raise

            if row_count:
                normalise_output_fonts(
                    output_file,
                    output_config,
                    last_data_row=row_count + 1,
                )
            LOGGER.info(
                "Completed %s: %d row(s) written in %.2f seconds",
                step_config.get("name", step_id),
                row_count,
                perf_counter() - step_started_at,
            )
            step_results.append(
                {
                    "step": step_config.get("name", step_id),
                    "rows": row_count,
                    "file": str(output_file.name),
                }
            )

        run_summary.update(
            {
                "application": application.get("name", "Spreadsheet ETL"),
                "entity_code": entity_code,
                "steps": step_results,
                "total_rows_written": sum(item["rows"] for item in step_results),
                "elapsed_seconds": round(perf_counter() - run_started_at, 2),
                "audit_log": log_path.name,
            }
        )
        summary_path = write_run_summary(run_summary, run_output_folder)
        LOGGER.info(
            "All configured steps completed in %.2f seconds",
            perf_counter() - run_started_at,
        )
        LOGGER.info("Run summary saved to %s", summary_path)
        LOGGER.info("Audit log saved to %s", log_path)
        return run_summary
    except Exception:
        LOGGER.exception("ETL run stopped")
        mark_failed_log(log_path)
        raise


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Run the generic spreadsheet ETL workflow."
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG_PATH,
        help="Path to a JSON configuration file (default: config.json)",
    )
    parser.add_argument(
        "--entity",
        help="Entity code to process; otherwise config or console selection is used.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_arguments()
    try:
        main(arguments.config, requested_entity=arguments.entity)
    except Exception as error:
        if not LOGGER.handlers:
            print(f"ETL run stopped: {error}", file=sys.stderr)
        raise SystemExit(1) from error
